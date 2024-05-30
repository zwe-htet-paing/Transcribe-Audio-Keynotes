import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

data = {
    "Audio": "/hom/kasdfadf/test.mp4",
    "Language": "english",
    "speaker": ["speaker_1", "speaker_2"],
    "Timestamp": ["2021-09-01 12:00:00", "2021-09-01 12:00:10"],
    "Transcript": ["Hello, how are you doing today?", "I am fine."],
    "task": "Transcribe",
}
keynote_data = {
    "Audio": "/hom/kasdfadf/test.mp4",
    "Language": "english",
    "KeyNotes": "Hello, how are you doing today? I am okay thank you. What are you doing now? Tell me about yourself. What is your occupation?",
    "task": "Note",
}


def json_to_excel(data, filename="audio_keynotes.xlsx"):
    """
    Converts a dictionary with potential list values to an Excel file and applies text wrapping if the task is 'Note'.

    Parameters:
        data (dict): The dictionary containing keys and values, where values can be lists or single elements.
        filename (str): The name of the Excel file to be saved.
    """
    # Create a list of dictionaries to be turned into a DataFrame
    data_for_df = []
    if data.get("task") == "Transcribe":
        max_len = max(
            len(value) if isinstance(value, list) else 1 for value in data.values()
        )
        for i in range(max_len):
            row_data = {
                key: (
                    data[key][i]
                    if isinstance(data[key], list) and i < len(data[key])
                    else data[key]
                )
                for key in data
            }
            data_for_df.append(row_data)
    else:
        row_data = {key: data[key] for key in data}
        data_for_df.append(row_data)
    # Create DataFrame
    df = pd.DataFrame(data_for_df)
    # Use openpyxl to create an Excel file with text wrapping for 'Note'
    wb = Workbook()
    ws = wb.active
    # Add DataFrame to Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    # Save the workbook
    wb.save(filename)
    print("Excel file has been created.")


if __name__ == "__main__":
    try:
        # json_to_excel(data)
        json_to_excel(keynote_data)
    except Exception as e:
        print("An error occurred:", e)
